import openpyxl


def create_xls(columns, data, file_name="signals.xlsx", translate_columns=False):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Словарь перевода колонок на русский язык
    column_translations = {
        "id": "ID",
        "user_id": "ID пользователя",
        "symbol": "Символ",
        "interval": "Интервал",
        "side": "Сторона",
        "qty": "Количество",
        "coin_buy_price": "Цена покупки монеты",
        "coin_sale_price": "Цена продажи монеты",
        "tp_price": "Цена Take Profit",
        "sl_price": "Цена Stop Loss",
        "buy_time": "Время покупки",
        "sale_time": "Время продажи",
        "status": "Статус",
        "pnl_usdt": "Прибыль/убыток в USDT",
        "investment_amount_usdt": "Сумма инвестиций в USDT",
        "pnl_percent": "Прибыль/убыток в процентах",
        "tp_percent": "Take Profit в процентах",
        "sl_percent": "Stop Loss в процентах",
        "return_amount_usdt": "Сумма возврата в USDT",
        "trading_type": "Тип торговли",
        "leverage": "Кредитное плечо",
        "exchange": "Биржа",
        "buy_price": "Цена покупки",
        "sale_price": "Цена продажи",
        "create_at": "Дата создания"
    }

    # Запись заголовков (переведенных или оригинальных)
    header_row = [column_translations.get(col, col) if translate_columns else col for col in columns]
    for col_num, column_name in enumerate(header_row, start=1):
        sheet.cell(row=1, column=col_num, value=column_name)

    # Запись данных
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num, value=cell_value)

    workbook.save(file_name)
    return file_name

def create_xls_stat(columns, data, file_name="orders_stat.xlsx"):
    return create_xls(columns, data, file_name, translate_columns=True)
